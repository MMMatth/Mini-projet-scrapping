## import the module ##
from openpyxl import load_workbook, Workbook
import os , requests, shutil

## function ##
# Fonction qui permet de modifier une case d'un fichier excel
def modif_cellule(fichier, feuille, ligne, colonne, valeur):
    """
    modif_cellule est une fonction qui permet de modifier une case d'un fichier excel

    Args:
        fichier (str): lien du fichier excel
        feuille (str): nom de la feuille
        ligne (int): numéro de la ligne
        colonne (int): numéro de la colonne
        valeur (_type_): valeur à mettre dans la case
    """
    wb = load_workbook(fichier) # on charge le fichier excel
    ws = wb[feuille] # on charge la feuille du fichier excel
    ws.cell(row=ligne, column=colonne).value = valeur # on modifie la case
    wb.save(fichier) # on sauvegarde le fichier excel

## class ##
# Class qui permet de récupérer les liens des pdf sur le site du prof et de les télécharger
class pdf:
    def __init__(self, lien : str, nom : str):
        self.nom = nom # nom du prof
        self.lien = lien  # lien du site du prof
        self.page = requests.get(self.lien)   # page du site du prof   
        
        self.texte = self.page.text.split("\n") # texte de la page du site du prof    
        self.nbr_dl = 0 # nombre de pdf téléchargé
        
        self.pdf = []  # liste des liens pdf
        self.recup_pdf() # on récupère les liens pdf
    

    def recup_pdf(self):
        """recup_pdf est une méthode qui permet de récupérer les liens pdf sur le site du prof"""
        for ligne in self.texte:
            if "pdf" in ligne: # on cherche les lignes qui contiennent le mot pdf
                if "http" in ligne: # si c'est une url
                    self.pdf.append(ligne[ligne.find("href=")+6 : ligne.find(".pdf")+4]) # on ajoute le lien du pdf dans la liste extrait de la ligne
                else: # sinon si c'est un lien relatif
                    self.pdf.append(self.lien[ :self.lien.rfind("/") ] # on enlève le nom du fichier du lien pour avoir que l'url
                                    + "/" + ligne[ ligne.find("href=")+6 : ligne.find(".pdf")+4 ]) # on ajoute le lien relatif a la suite de l'url
    
    def dl_pdf(self, path : str):
        """
        dl_pdf est une méthode qui permet de télécharger les pdf dans un dossier

        Args:
            path (str): chemin du dossier où télécharger les pdf
        """
        for lien in range(0,len(self.pdf)): 
            r = requests.get(self.pdf[lien], stream=True) # on récupère le pdf     
            with open(path + str( lien + 1 ) + ".pdf" , 'wb') as f: # on le télécharge dans le dossier
                f.write(r.content) # on écrit le contenu du pdf dans le fichier 
            self.nbr_dl += 1 # on incrémente le nombre de pdf télécharger
         

## main ##
## lire le fichier excel et mettre les données dans un dictionnaire ##
wb = load_workbook(filename = 'liste_ens.xlsx')
sheet_ranges = wb['Feuil1']
site = { # dictionnaire qui contient les noms des profs comme clé, le lien comme premier élément de la liste et l'objet pdf comme deuxième élément de la liste ( None pour l'instant )
    sheet_ranges['B2'].value: [sheet_ranges['C2'].value, None],
    sheet_ranges['B3'].value: [sheet_ranges['C3'].value, None],
    sheet_ranges['B4'].value: [sheet_ranges['C4'].value, None],
}

## on suprime les anciens fichiers (sert pendant les testes mais jamais au premier lancement)##
# for prenom in site.keys():
#     shutil.rmtree(prenom)

## crée un dossier par professeur ##
for prenom in site.keys(): 
    os.mkdir(prenom)
    
## on récupère les liens des pdf sur les sites ##
l_table = 2 # numéro de la ligne dans le fichier excel
for prenom in site.keys(): # pour chaque prof
    site[prenom][1] = pdf(site[prenom][0],prenom) # on crée un objet pdf pour chaque prof
    site[prenom][1].dl_pdf(str(prenom)+"/"+str(prenom)) # on télécharge les pdf dans le dossier de chaque prof
    modif_cellule("liste_ens.xlsx", "Feuil1", l_table , 4, site[prenom][1].nbr_dl) # on modifie le fichier excel
    l_table += 1 # on incrémente l_table pour aller à la ligne suivante du fichier excel
    
